import openpyxl
import matplotlib.pyplot as plt
import japanize_matplotlib

wb = openpyxl.load_workbook("陽性者数.xlsx")
sheet = wb["陽性者数(2021)"]

result = list()
count_list = list()
count = 0

for i in range (184, 215):
    count = count + sheet.cell(row = i, column = 2).value
    result.append(sheet.cell(row = i, column = 2).value)
    count_list.append(count)

plt.plot(result)
plt.plot(count_list)
plt.title("2021年10月の陽性者数と累計陽性者の推移")
plt.xlabel('日付')
plt.ylabel('陽性者数[人]')
plt.legend(['陽性者数', '累計陽性者数'])
plt.xticks([0, 7, 14, 21, 28], ['10/1', '10/8', '10/15', '10/22', '10/31'])
plt.grid()
plt.savefig('covid_2021_Oct.svg')
plt.show()

wb.close()