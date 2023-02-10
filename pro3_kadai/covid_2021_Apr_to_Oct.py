import openpyxl
import matplotlib.pyplot as plt
import japanize_matplotlib

wb = openpyxl.load_workbook("陽性者数.xlsx")
sheet = wb["陽性者数(2021)"]

result = list()

count = 0
for i in range (1, 31):
    count = count + sheet.cell(row = i, column = 2).value
result.append(count)

count = 0
for i in range (31, 62):
    count = count + sheet.cell(row = i, column = 2).value
result.append(count)

count = 0
for i in range (62, 92):
    count = count + sheet.cell(row = i, column = 2).value
result.append(count)

count = 0
for i in range (92, 123):
    count = count + sheet.cell(row = i, column = 2).value
result.append(count)

count = 0
for i in range (123, 154):
    count = count + sheet.cell(row = i, column = 2).value
result.append(count)

count = 0
for i in range (154, 184):
    count = count + sheet.cell(row = i, column = 2).value
result.append(count)

count = 0
for i in range (184, 215):
    count = count + sheet.cell(row = i, column = 2).value
result.append(count)

plt.plot(result)
plt.title("2021年4月から10月までの陽性者の推移")
plt.xlabel('月')
plt.ylabel('陽性者数[人]')
plt.xticks([0, 1, 2, 3, 4, 5, 6], ['4月', '5月', '6月', '7月', '8月', '9月', '10月'])
plt.grid()
plt.savefig('covid_2021_Apr_to_Oct.svg')
plt.show()

wb.close()