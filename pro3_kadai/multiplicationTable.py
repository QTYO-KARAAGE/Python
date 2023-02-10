import openpyxl as excel
import sys

N = int(sys.argv[1])

#Excelファイルの読み込み
wb = excel.load_workbook("multiplicationTable.xlsx")
sheet = wb.active

#テーブルの横軸作成
for i in range(1, N + 1):
    c = sheet.cell(1, i + 1)
    c.value = i
#テーブルの縦軸作成
for j in range(1, N + 1):
    c = sheet.cell(j + 1, 1)
    c.value = j

#値の計算と代入
for i in range(1, N + 1):
    for j in range(1, N + 1):
        ans = sheet.cell(i + 1, j + 1)
        row = sheet.cell(1, j + 1)
        column = sheet.cell(i + 1, 1)
        ans.value = row.value * column.value

#保存と終了
wb.save('multiplicationTable.xlsx')
wb.close()